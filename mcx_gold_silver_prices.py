import pandas as pd
import requests
from datetime import datetime, timedelta
import json
import time

def fetch_silver_data_demo():
    """
    Demo function to create sample MCX Silver futures data
    This simulates the data structure you would get from Upstox API
    """
    
    # Generate sample dates from Jan 1, 2025 to current date
    start_date = datetime(2023, 1, 1)
    end_date = datetime.now()
    
    date_range = pd.date_range(start=start_date, end=end_date, freq='D')
    
    # Create sample silver futures data (realistic price movements)
    import random
    random.seed(42)  # For reproducible results
    
    base_price = 105000  # Starting price around 	2,05,000
    data = []
    
    for i, date in enumerate(date_range):
        # Skip weekends (Saturday=5, Sunday=6)
        if date.weekday() >= 5:
            continue
            
        # Simulate price movement
        change = random.uniform(-2000, 2000)  # Daily change up to 	2,000
        base_price += change
        base_price = max(base_price, 95000)  # Floor price
        base_price = min(base_price, 115000)  # Ceiling price
        
        # Generate OHLC data
        open_price = base_price + random.uniform(-500, 500)
        close_price = base_price + random.uniform(-500, 500)
        high_price = max(open_price, close_price) + random.uniform(0, 1000)
        low_price = min(open_price, close_price) - random.uniform(0, 1000)
        
        volume = random.randint(1000, 10000)
        open_interest = random.randint(5000, 50000)
        
        data.append({
            'Date': date.strftime('%Y-%m-%d'),
            'Open': round(open_price, 2),
            'High': round(high_price, 2),
            'Low': round(low_price, 2),
            'Close': round(close_price, 2),
            'Volume': volume,
            'Open_Interest': open_interest,
            'Change': round(close_price - open_price, 2),
            'Change_%': round(((close_price - open_price) / open_price) * 100, 2)
        })
    
    return pd.DataFrame(data)

def fetch_gold_data_demo():
    """
    Demo function to create sample MCX Gold futures data
    This simulates the data structure you would get from Upstox API
    """
    
    # Generate sample dates from Jan 1, 2025 to current date
    start_date = datetime(2025, 1, 1)
    end_date = datetime.now()
    
    date_range = pd.date_range(start=start_date, end=end_date, freq='D')
    
    # Create sample gold futures data (realistic price movements)
    import random
    random.seed(24)  # Different seed for gold
    
    base_price = 600000  # Starting price around 	2,00,000
    data = []
    
    for i, date in enumerate(date_range):
        # Skip weekends (Saturday=5, Sunday=6)
        if date.weekday() >= 5:
            continue
            
        # Simulate price movement
        change = random.uniform(-5000, 5000)  # Daily change up to 	2,000
        base_price += change
        base_price = max(base_price, 550000)  # Floor price
        base_price = min(base_price, 650000)  # Ceiling price
        
        # Generate OHLC data
        open_price = base_price + random.uniform(-1000, 1000)
        close_price = base_price + random.uniform(-1000, 1000)
        high_price = max(open_price, close_price) + random.uniform(0, 2000)
        low_price = min(open_price, close_price) - random.uniform(0, 2000)
        
        volume = random.randint(500, 5000)
        open_interest = random.randint(2000, 20000)
        
        data.append({
            'Date': date.strftime('%Y-%m-%d'),
            'Open': round(open_price, 2),
            'High': round(high_price, 2),
            'Low': round(low_price, 2),
            'Close': round(close_price, 2),
            'Volume': volume,
            'Open_Interest': open_interest,
            'Change': round(close_price - open_price, 2),
            'Change_%': round(((close_price - open_price) / open_price) * 100, 2)
        })
    
    return pd.DataFrame(data)

def export_to_excel_advanced(df_silver, df_gold, filename='filtered_data_Jul.xlsx'):
    """
    Export Silver and Gold DataFrames to Excel with advanced formatting
    """
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write silver data
            df_silver.to_excel(writer, sheet_name='Silver Futures', index=False, startrow=3)
            # Write gold data
            df_gold.to_excel(writer, sheet_name='Gold Futures', index=False, startrow=3)
            
            # Get workbook and worksheets
            workbook = writer.book
            ws_silver = writer.sheets['Silver Futures']
            ws_gold = writer.sheets['Gold Futures']
            
            # Add titles and metadata
            ws_silver['A1'] = 'MCX Silver Futures Contract Data'
            ws_silver['A2'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws_silver['A3'] = f'Data Period: {df_silver["Date"].min()} to {df_silver["Date"].max()}'
            
            ws_gold['A1'] = 'MCX Gold Futures Contract Data'
            ws_gold['A2'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws_gold['A3'] = f'Data Period: {df_gold["Date"].min()} to {df_gold["Date"].max()}'
            
            # Format titles
            from openpyxl.styles import Font, PatternFill, Alignment
            title_font = Font(size=16, bold=True)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            ws_silver['A1'].font = title_font
            ws_gold['A1'].font = title_font
            
            # Format headers (row 4) for both sheets
            for ws in [ws_silver, ws_gold]:
                for col in range(1, len(df_silver.columns) + 1):
                    cell = ws.cell(row=4, column=col)
                    cell.font = Font(color='FFFFFF', bold=True)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 20)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"✅ Data successfully exported to {filename}")
        return True
        
    except Exception as e:
        print(f"❌ Error exporting to Excel: {e}")
        return False

def display_data_preview(df, commodity_name):
    """
    Display a nice preview of the data
    """
    print("=" * 80)
    print(f"📊 MCX {commodity_name.upper()} FUTURES DATA PREVIEW")
    print("=" * 80)
    
    print(f"\n📅 Data Period: {df['Date'].min()} to {df['Date'].max()}")
    print(f"📈 Total Trading Days: {len(df)}")
    
    print("\n🔝 FIRST 5 RECORDS:")
    print(df.head().to_string(index=False))
    
    print("\n🔚 LAST 5 RECORDS:")
    print(df.tail().to_string(index=False))
    
    print("\n📊 SUMMARY STATISTICS:")
    print(f"💰 Highest Price: ₹{df['High'].max():,.2f}")
    print(f"💰 Lowest Price: ₹{df['Low'].min():,.2f}")
    print(f"💰 Average Close: ₹{df['Close'].mean():,.2f}")
    print(f"📊 Total Volume: {df['Volume'].sum():,}")
    print(f"📈 Best Day: +₹{df['Change'].max():,.2f}")
    print(f"📉 Worst Day: ₹{df['Change'].min():,.2f}")
    
    print("\n" + "=" * 80)

def main():
    """
    Main function - Ready to run!
    """
    print("🚀 Starting MCX Silver and Gold Futures Data Extraction...")
    print("⏳ Generating sample data (simulating API call)...")
    
    # Simulate loading time
    time.sleep(2)
    
    # Fetch the data
    df_silver = fetch_silver_data_demo()
    df_gold = fetch_gold_data_demo()
    
    if df_silver is not None and not df_silver.empty and df_gold is not None and not df_gold.empty:
        print(f"✅ Successfully generated {len(df_silver)} silver records and {len(df_gold)} gold records")
        
        # Display preview
        display_data_preview(df_silver, 'silver')
        display_data_preview(df_gold, 'gold')
        
        # Export to Excel
        filename = f'filtered_data_Jul{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        print(f"\n💾 Exporting data to Excel: {filename}")
        success = export_to_excel_advanced(df_silver, df_gold, filename)
        
        if success:
            print(f"\n🎉 SUCCESS! Your MCX Silver and Gold futures data is ready!")
            print(f"📁 File saved as: {filename}")
            print(f"📊 Silver data trading days: {len(df_silver)}")
            print(f"📊 Gold data trading days: {len(df_gold)}")
            print(f"📈 Date range: {df_silver['Date'].min()} to {df_silver['Date'].max()} (Silver)")
            print(f"📈 Date range: {df_gold['Date'].min()} to {df_gold['Date'].max()} (Gold)")
        else:
            print("❌ Failed to export data")
    else:
        print("❌ Failed to generate data")

# Additional utility function to install required packages
# def install_requirements():
#     """
#     Install required packages if not available
#     """
#     required_packages = ['pandas', 'openpyxl']
#     
#     for package in required_packages:
#         try:
#             __import__(package)
#         except ImportError:
#             print(f"Installing {package}...")
#             import subprocess
#             subprocess.check_call(['pip', 'install', package])

if __name__ == "__main__":
    # Uncomment the line below if you need to install packages
    # install_requirements()
    
    main()
